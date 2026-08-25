from openpyxl import load_workbook


def extract_content_from_excel(file_path):
    workbook_formulas = load_workbook(
        file_path,
        data_only=False
    )

    workbook_values = load_workbook(
        file_path,
        data_only=True
    )

    blocks = []

    for sheet_name in workbook_formulas.sheetnames:
        formula_sheet = workbook_formulas[sheet_name]
        value_sheet = workbook_values[sheet_name]

        for row_number, formula_row in enumerate(
            formula_sheet.iter_rows(),
            start=1
        ):
            row_parts = []

            for column_number, formula_cell in enumerate(
                formula_row,
                start=1
            ):
                formula_value = formula_cell.value

                if formula_value is None:
                    continue

                value_cell = value_sheet.cell(
                    row=row_number,
                    column=column_number
                )

                calculated_value = value_cell.value

                if (
                    isinstance(formula_value, str)
                    and formula_value.startswith("=")
                ):
                    row_parts.append(
                        f"{formula_cell.coordinate}: "
                        f"Formula={formula_value}, "
                        f"Value={calculated_value}"
                    )
                else:
                    row_parts.append(
                        f"{formula_cell.coordinate}: "
                        f"{formula_value}"
                    )

            if row_parts:
                blocks.append({
                    "type": "table",
                    "content": " | ".join(row_parts),
                    "location": (
                        f"{sheet_name} - Row {row_number}"
                    ),
                    "metadata": {
                        "sheet": sheet_name,
                        "row": row_number
                    }
                })

    return blocks